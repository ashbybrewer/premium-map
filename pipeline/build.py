#!/usr/bin/env python3
"""
THE PREMIUM MAP — data pipeline
Transforms raw public datasets into the compact JS data modules the instrument loads.

Inputs (see pipeline/fetch.sh):
  raw/fio-metrics.xlsx      Treasury FIO / NAIC PCMI ZIP-level metrics, 2018-2022  [SOURCED]
  raw/senate-nonrenewal.xlsx Senate Budget Cmte county non-renewal data, 2018-2023 [SOURCED]
  raw/zcta-county.txt       Census 2020 ZCTA -> county relationship file           [SOURCED]
  raw/fema-decl.csv         OpenFEMA disaster declarations (county), 2000+         [SOURCED]
  raw/noaa-events.csv       NOAA/NCEI billion-dollar disasters, 1980-2024 (final)  [SOURCED]
  raw/counties-10m.json     us-atlas TopoJSON

Outputs:
  data/counties.js, data/national.js, data/topology.js
  data/processed/county_metrics.csv
Every MODELED figure is produced only here, flagged, and documented in README.
"""
import csv, json, math, re, sys, os
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
RAW  = os.path.join(ROOT, 'raw')
OUT  = os.path.join(ROOT, 'data')
os.makedirs(os.path.join(OUT, 'processed'), exist_ok=True)

STATE_FIPS = {
 '01':'AL','02':'AK','04':'AZ','05':'AR','06':'CA','08':'CO','09':'CT','10':'DE','11':'DC',
 '12':'FL','13':'GA','15':'HI','16':'ID','17':'IL','18':'IN','19':'IA','20':'KS','21':'KY',
 '22':'LA','23':'ME','24':'MD','25':'MA','26':'MI','27':'MN','28':'MS','29':'MO','30':'MT',
 '31':'NE','32':'NV','33':'NH','34':'NJ','35':'NM','36':'NY','37':'NC','38':'ND','39':'OH',
 '40':'OK','41':'OR','42':'PA','44':'RI','45':'SC','46':'SD','47':'TN','48':'TX','49':'UT',
 '50':'VT','51':'VA','53':'WA','54':'WV','55':'WI','56':'WY'}
ABBR_FIPS = {v:k for k,v in STATE_FIPS.items()}

# MODELED nowcast anchors: S&P Global Market Intelligence RateWatch national
# weighted-average approved rate changes (owner-occupied HO): 2023 +12.7%, 2024 +10.4%,
# 2025 +6.0% (partial-year reporting). Applied uniformly to every county's 2022 premium.
ANCHORS = {'2023': 1.127, '2024': 1.104, '2025': 1.060}

import unicodedata
def norm_name(s):
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.lower().strip()
    s = re.sub(r'[.\'’]', '', s)
    s = re.sub(r'\s+', ' ', s)
    for suf in [' county',' city and borough',' borough',' census area',
                ' municipality',' municipio',' parish']:
        if s.endswith(suf): s = s[:-len(suf)]
    s = s.replace('saint ','st ').replace('sainte ','ste ')
    s = re.sub(r'^(la|de) ?salle$', 'lasalle', s)
    return s.strip()

# ---------------------------------------------------------------- crosswalk
print('== ZCTA -> county crosswalk')
best = {}   # zcta -> (area, county_geoid)
county_name = {}  # geoid -> NAMELSAD
with open(os.path.join(RAW,'zcta-county.txt'), encoding='utf-8-sig') as f:
    rdr = csv.DictReader(f, delimiter='|')
    for r in rdr:
        g = r['GEOID_COUNTY_20']
        if g and g[:2] in STATE_FIPS:
            county_name[g] = r['NAMELSAD_COUNTY_20']
        z = r['GEOID_ZCTA5_20']
        if not z or not g: continue
        try: a = int(r['AREALAND_PART'] or 0)
        except: a = 0
        if z not in best or a > best[z][0]:
            best[z] = (a, g)
zip2cty = {z: g for z,(a,g) in best.items()}
print('  zctas mapped:', len(zip2cty), '| counties named:', len(county_name))

# name -> fips lookup for Senate matching
name2fips = {}
for g, nm in county_name.items():
    st = STATE_FIPS.get(g[:2])
    if not st: continue
    name2fips[(st, norm_name(nm))] = g
    # independent cities: also index with ' city' retained distinctly
    if nm.lower().endswith(' city'):
        name2fips[(st, norm_name(nm[:-5]) + ' city')] = g

# ---------------------------------------------------------------- FIO PCMI
print('== FIO PCMI ZIP-level metrics (2018-2022)')
import openpyxl
wb = openpyxl.load_workbook(os.path.join(RAW,'fio-metrics.xlsx'), read_only=True)
ws = wb['Supporting Underlying Metrics']
YEARS = [2018,2019,2020,2021,2022]
acc  = defaultdict(lambda: defaultdict(lambda: [0.0]*10))  # county -> year -> sums
nat  = defaultdict(lambda: [0.0]*10)                       # year -> sums
unmatched_zip = set()
nrows = 0
def fnum(v):
    if v is None or v == '' : return None
    try:
        x = float(v)
        return x if math.isfinite(x) else None
    except: return None
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    z, yr, dec, cf, cs, lr, prem, nr, npc, onpc = row[:10]
    if z is None or yr is None: continue
    z = str(z).split('.')[0].zfill(5)
    try: yr = int(yr)
    except: continue
    if yr not in (2018,2019,2020,2021,2022): continue
    w = fnum(dec) or 1.0
    cty = zip2cty.get(z)
    if cty is None:
        unmatched_zip.add(z); continue
    prem, nr, lr, cf, cs = fnum(prem), fnum(nr), fnum(lr), fnum(cf), fnum(cs)
    for tgt in (acc[cty][yr], nat[yr]):
        if prem is not None: tgt[0]+=w*prem; tgt[1]+=w
        if nr   is not None: tgt[2]+=w*nr;   tgt[3]+=w
        if lr   is not None: tgt[4]+=w*min(lr,5.0); tgt[5]+=w   # clamp pathological loss ratios
        if cf   is not None: tgt[6]+=w*cf;   tgt[7]+=w
        if cs   is not None: tgt[8]+=w*cs;   tgt[9]+=w
    nrows += 1
print(f'  rows used: {nrows} | zips w/o county match: {len(unmatched_zip)}')

def wavg(s, i): return (s[i]/s[i+1]) if s[i+1] > 0 else None

fio = {}
for cty, ys in acc.items():
    prem = [wavg(ys[y],0) for y in YEARS]
    fio[cty] = {
        'p':  prem,
        'fr': [wavg(ys[y],2) for y in YEARS],
        'lr': [wavg(ys[y],4) for y in YEARS],
        'cf': [wavg(ys[y],6) for y in YEARS],
        'cs': [wavg(ys[y],8) for y in YEARS],
    }
nat_prem = [wavg(nat[y],0) for y in YEARS]
nat_fr   = [wavg(nat[y],2) for y in YEARS]
nat_lr   = [wavg(nat[y],4) for y in YEARS]
nat_cf   = [wavg(nat[y],6) for y in YEARS]
print('  national premium/policy 2018..2022:', [round(x,1) for x in nat_prem])
print('  national FIO nonrenewal 2018..2022:', [round(x,4) for x in nat_fr])

# ---------------------------------------------------------------- Senate non-renewals
print('== Senate Budget Committee county non-renewals (2018-2023)')
wb2 = openpyxl.load_workbook(os.path.join(RAW,'senate-nonrenewal.xlsx'), read_only=True)
SYEARS = [2018,2019,2020,2021,2022,2023]
sen = defaultdict(lambda: {'nr':[None]*6, 'pol':[None]*6})
sen_nat_nr = [0.0]*6; sen_nat_pol = [0.0]*6
unmatched_names = []
nospace2fips = {(st, k.replace(' ','')): g for (st,k),g in name2fips.items()}
def lookup(st, raw):
    k = norm_name(raw)
    g = name2fips.get((st, k))
    if g is None: g = name2fips.get((st, k + ' city'))
    if g is None and k.endswith(' city'):
        g = name2fips.get((st, k[:-5].strip()))
    if g is None: g = nospace2fips.get((st, k.replace(' ','')))
    return g
SKIP_LABELS = {'UNKNOWN','GRAND TOTAL','STATEWIDE','OTHER','TOTAL'}
ALIASES = {('CA','SAN BERNANDINO'):'06071', ('CA','ANGELES'):'06037',
           ('GA','MACON-BIBB'):'13021'}  # source typos / consolidated gov names

for sy_i, y in enumerate(SYEARS):
    ws2 = wb2[f'{y} Compiled']
    cur_st = None
    for row in ws2.iter_rows(min_row=2, values_only=True):
        label, n_nr, pol_eoy, pol_fy, rate = row[:5]
        if label is None: continue
        label = str(label).strip()
        if re.fullmatch(r'[A-Za-z]{2}', label) and label.upper() in ABBR_FIPS:
            cur_st = label.upper(); continue
        if cur_st is None or label.upper() in SKIP_LABELS: continue
        rate_f = fnum(rate); pol_f = fnum(pol_fy); nr_f = fnum(n_nr)
        if nr_f is not None and pol_f:                    # national sums once per row
            sen_nat_nr[sy_i] += nr_f; sen_nat_pol[sy_i] += pol_f
        parts = [p.strip() for p in label.split('/')] if '/' in label else [label]
        combo = len(parts) > 1
        for part in parts:
            g = ALIASES.get((cur_st, part.upper())) or lookup(cur_st, part)
            if g is None:
                unmatched_names.append(f'{y}|{cur_st}|{part}')
                continue
            if rate_f is not None: sen[g]['nr'][sy_i] = rate_f
            if pol_f is not None and not combo: sen[g]['pol'][sy_i] = pol_f
nat_sen_nr = [(sen_nat_nr[i]/sen_nat_pol[i]) if sen_nat_pol[i] else None for i in range(6)]
print(f'  counties matched: {len(sen)} | unmatched rows: {len(unmatched_names)}')
print('  sample unmatched:', unmatched_names[:10])
print('  national senate nonrenewal 2018..2023:', [None if x is None else round(x,4) for x in nat_sen_nr])

# ---------------------------------------------------------------- FEMA declarations
print('== OpenFEMA disaster declarations (2000+)')
CLIMATE = {'Hurricane':'hur','Typhoon':'hur','Fire':'fire','Severe Storm':'storm',
 'Severe Storm(s)':'storm','Tornado':'storm','Coastal Storm':'storm','Severe Ice Storm':'winter',
 'Snowstorm':'winter','Winter Storm':'winter','Freezing':'winter','Flood':'flood','Dam/Levee Break':'flood',
 'Mud/Landslide':'flood','Tropical Storm':'hur','Drought':'other','Earthquake':None,'Biological':None,
 'Chemical':None,'Terrorist':None,'Other':None,'Toxic Substances':None,'Human Cause':None,
 'Fishing Losses':None,'Volcanic Eruption':None,'Tsunami':None,'Straight-Line Winds':'storm','Winter':'winter'}
GKEYS = ['hur','fire','storm','flood','winter','other']
decl = defaultdict(lambda: {'t':0,'g':[0]*6,'l5':0,'ls':[]})
seen = set()
with open(os.path.join(RAW,'fema-decl.csv'), encoding='utf-8') as f:
    rdr = csv.DictReader(f)
    for r in rdr:
        sc, cc = r['fipsStateCode'], r['fipsCountyCode']
        if not sc or not cc or cc == '000': continue
        if sc not in STATE_FIPS: continue
        g = sc + cc
        grp = CLIMATE.get(r['incidentType'])
        if grp is None: continue
        y = int(r['declarationDate'][:4])
        key = (g, r['disasterNumber'])
        if key in seen: continue
        seen.add(key)
        d = decl[g]
        d['t'] += 1
        d['g'][GKEYS.index(grp)] += 1
        if y >= 2021: d['l5'] += 1
        title = r['declarationTitle'].title()[:44]
        d['ls'].append((y, grp, title))
for g,d in decl.items():
    d['ls'] = sorted(d['ls'], key=lambda x:-x[0])[:5]
print('  counties with declarations:', len(decl))

# ---------------------------------------------------------------- NOAA billion-$ events
print('== NOAA billion-dollar events (1980-2024, final release)')
noaa_year = defaultdict(lambda: {'c':0.0,'n':0,'d':0})
noaa_top = []
with open(os.path.join(RAW,'noaa-events.csv'), encoding='utf-8') as f:
    lines = [ln for ln in f]
data_start = next(i for i,ln in enumerate(lines) if ln.startswith('Name,'))
rdr = csv.DictReader(lines[data_start:])
for r in rdr:
    try:
        y = int(str(r['Begin Date'])[:4])
        cost = float(r['CPI-Adjusted Cost'])
        deaths = int(float(r.get('Deaths') or 0))
    except: continue
    noaa_year[y]['c'] += cost; noaa_year[y]['n'] += 1; noaa_year[y]['d'] += deaths
    noaa_top.append({'name': r['Name'], 'y': y, 'cost': round(cost/1000.0,1),
                     'type': r['Disaster'], 'deaths': deaths})
noaa_top = sorted(noaa_top, key=lambda e:-e['cost'])[:15]
noaa_series = [{'y':y, 'c': round(noaa_year[y]['c']/1000.0,1), 'n': noaa_year[y]['n']}
               for y in sorted(noaa_year)]
print('  years:', len(noaa_series), '| 2024:', noaa_series[-1])

# ---------------------------------------------------------------- assemble counties
print('== assemble')
r2 = lambda x, nd=2: (None if x is None else round(x, nd))
counties = {}
all_fips = set(fio) | set(sen) | set(decl)
for g in all_fips:
    nm = county_name.get(g)
    if nm is None: continue
    st = STATE_FIPS.get(g[:2])
    o = {'n': re.sub(r' (County|Parish)$','',nm), 'st': st}
    f = fio.get(g)
    if f and any(v is not None for v in f['p']):
        o['p']  = [r2(v,0) for v in f['p']]
        o['fr'] = [r2(v,4) for v in f['fr']]
        o['lr'] = [r2(v,3) for v in f['lr']]
        o['cf'] = [r2(v,4) for v in f['cf']]
        css = [v for v in f['cs'] if v is not None]
        o['cs'] = r2(sum(css)/len(css),0) if css else None
        if f['p'][4] is not None:
            base = f['p'][4]
            o['m'] = [r2(base*ANCHORS['2023'],0),
                      r2(base*ANCHORS['2023']*ANCHORS['2024'],0),
                      r2(base*ANCHORS['2023']*ANCHORS['2024']*ANCHORS['2025'],0)]
    s = sen.get(g)
    if s:
        o['nr']  = [r2(v,4) for v in s['nr']]
        o['pf']  = s['pol'][5] and int(s['pol'][5])
    d = decl.get(g)
    if d:
        o['dr'] = {'t': d['t'], 'g': d['g'], 'l5': d['l5'], 'ls': d['ls']}
    counties[g] = o

# growth & percentile ranks
def pct_ranks(pairs):  # [(fips, value)] -> fips -> percentile 0..100
    pairs = [(g,v) for g,v in pairs if v is not None]
    pairs.sort(key=lambda x:x[1])
    n = len(pairs)
    return {g: 100.0*i/(n-1) for i,(g,v) in enumerate(pairs)} if n>1 else {}

growth = {}
for g,o in counties.items():
    p = o.get('p')
    if p and p[0] and p[4]:
        growth[g] = p[4]/p[0] - 1.0
        o['gw'] = r2(growth[g]*100,1)
rk_g  = pct_ranks(list(growth.items()))
rk_p  = pct_ranks([(g, o['p'][4] if o.get('p') else None) for g,o in counties.items()])
rk_nr = pct_ranks([(g, o['nr'][5] if o.get('nr') else None) for g,o in counties.items()])
rk_lr = pct_ranks([(g, (sum(v for v in o['lr'] if v is not None)/max(1,sum(1 for v in o['lr'] if v is not None))) if o.get('lr') else None) for g,o in counties.items()])
rk_d  = pct_ranks([(g, o['dr']['l5'] if o.get('dr') else 0) for g,o in counties.items()])

# MODELED composite: Repricing Index 0-100
for g,o in counties.items():
    parts, wts = [], []
    if g in rk_g:  parts.append(rk_g[g]);  wts.append(0.35)
    if g in rk_nr: parts.append(rk_nr[g]); wts.append(0.25)
    if g in rk_lr: parts.append(rk_lr[g]); wts.append(0.20)
    if g in rk_d:  parts.append(rk_d[g]);  wts.append(0.20)
    if parts and sum(wts) >= 0.55:
        o['dx'] = r2(sum(p*w for p,w in zip(parts,wts))/sum(wts),1)
    rk = {}
    if g in rk_p:  rk['p']  = r2(rk_p[g],0)
    if g in rk_g:  rk['g']  = r2(rk_g[g],0)
    if g in rk_nr: rk['nr'] = r2(rk_nr[g],0)
    if g in rk_lr: rk['lr'] = r2(rk_lr[g],0)
    if rk: o['rk'] = rk

have_p = sum(1 for o in counties.values() if o.get('p'))
have_n = sum(1 for o in counties.values() if o.get('nr'))
print(f'  counties total: {len(counties)} | with premiums: {have_p} | with senate nr: {have_n}')

# ---------------------------------------------------------------- write outputs
nat_out = {
 'years': YEARS, 'senYears': SYEARS,
 'prem':  [r2(v,0) for v in nat_prem],
 'premE': [r2(nat_prem[4]*ANCHORS['2023'],0),
           r2(nat_prem[4]*ANCHORS['2023']*ANCHORS['2024'],0),
           r2(nat_prem[4]*ANCHORS['2023']*ANCHORS['2024']*ANCHORS['2025'],0)],
 'fr':  [r2(v,4) for v in nat_fr],
 'lr':  [r2(v,3) for v in nat_lr],
 'cf':  [r2(v,4) for v in nat_cf],
 'nr':  [r2(v,4) for v in nat_sen_nr],
 'anchors': {'2023': 12.7, '2024': 10.4, '2025': 6.0},
 'noaa': {'series': noaa_series, 'top': noaa_top,
          'note': 'NOAA/NCEI retired the Billion-Dollar Disasters product in May 2025; 1980-2024 is the final federal ledger.'},
 'exits': [
  {'d':'2021-07','st':'FL','co':'Gulfstream P&C','k':'insolvency','ev':'Liquidated July 28, 2021 — first of nine Florida property insurers to fail in the 2021–23 wave.','src':'https://figafacts.com/category/insolvency/'},
  {'d':'2021-08','st':'LA','co':'Eleven insurers','k':'insolvency','ev':'Hurricanes Laura & Ida trigger a cascade of insolvencies through Sept 2022 — ~185,000 policies, 13% of the Louisiana market.','src':'https://www.nola.com/news/business/how-failed-lousiana-insurers-left-homeowners-worse-off/article_0b37cf1e-b188-11ee-aa68-4fe643b958d8.html'},
  {'d':'2022-02','st':'FL','co':'St. Johns Insurance','k':'insolvency','ev':'Liquidated February 25, 2022.','src':'https://figafacts.com/category/insolvency/'},
  {'d':'2022-03','st':'FL','co':'Avatar P&C','k':'insolvency','ev':'Liquidated March 14, 2022.','src':'https://figafacts.com/category/insolvency/'},
  {'d':'2022-04','st':'FL','co':'Lighthouse Property','k':'insolvency','ev':'Liquidated April 28, 2022.','src':'https://figafacts.com/category/insolvency/'},
  {'d':'2022-06','st':'FL','co':'Southern Fidelity','k':'insolvency','ev':'Liquidated June 15, 2022.','src':'https://figafacts.com/category/insolvency/'},
  {'d':'2022-08','st':'FL','co':'Weston P&C','k':'insolvency','ev':'Liquidated August 8, 2022.','src':'https://figafacts.com/category/insolvency/'},
  {'d':'2022-09','st':'FL','co':'FedNat','k':'insolvency','ev':'Liquidated September 27, 2022.','src':'https://figafacts.com/category/insolvency/'},
  {'d':'2022-11','st':'CA','co':'Allstate','k':'pause','ev':'Quietly pauses new home & condo policies statewide to manage wildfire exposure (disclosed June 2023).','src':'https://www.sfexaminer.com/news/housing/farmers-csaa-not-leaving-ca-after-allstate-state-farm/article_4e4b94ac-0562-11ee-85bd-a34183a8f1ee.html'},
  {'d':'2023-02','st':'FL','co':'United P&C','k':'insolvency','ev':'Liquidated February 27, 2023 — the largest of Florida’s nine failures.','src':'https://www.tampabay.com/news/florida-politics/2023/04/17/homeowners-insurance-property-insolvent-rising-rates/'},
  {'d':'2023-05','st':'CA','co':'State Farm','k':'pause','ev':'Halts ALL new property applications in California, effective May 27, 2023.','src':'https://www.insurance.com/home-insurance/state-farm-stops-selling-home-insurance-in-california'},
  {'d':'2023-07','st':'FL','co':'Farmers','k':'exit','ev':'Withdraws from Florida — roughly 100,000 policyholders affected.','src':'https://www.insurancebusinessmag.com/us/news/property/major-insurance-company-abandons-homeowners-in-two-key-states-452325.aspx'},
  {'d':'2023-07','st':'CA','co':'Farmers','k':'pause','ev':'Caps new California policies at ~7,000 per month.','src':'https://www.sfexaminer.com/news/housing/farmers-csaa-not-leaving-ca-after-allstate-state-farm/article_4e4b94ac-0562-11ee-85bd-a34183a8f1ee.html'},
  {'d':'2023-09','st':'NC','co':'Nationwide','k':'exit','ev':'Nonrenews 10,525 coastal North Carolina policies east of I-95, citing hurricane risk.','src':'https://www.insurancejournal.com/news/southeast/2023/09/29/742310.htm'},
  {'d':'2024-02','st':'CA','co':'The Hartford','k':'exit','ev':'Stops writing new California homeowners policies from February 1, 2024.','src':'https://www.insurancejournal.com/news/west/2024/01/24/757094.htm'},
  {'d':'2024-03','st':'CA','co':'State Farm','k':'exit','ev':'Announces nonrenewal of ~30,000 home policies (72,000 incl. commercial), concentrated in high-wildfire ZIP codes.','src':'https://www.insurancejournal.com/news/west/2024/03/20/765883.htm'},
  {'d':'2024-04','st':'CA','co':'Tokio Marine + Trans Pacific','k':'exit','ev':'Both withdraw from California homeowners entirely — 12,556 policies, $11.3M premium.','src':'https://www.insurancejournal.com/news/west/2024/04/18/770680.htm'},
  {'d':'2025-01','st':'CA','co':'State Farm','k':'pause','ev':'After the Los Angeles fires, pauses pending nonrenewals in compliance with CDI notice.','src':'https://www.carriermanagement.com/news/2025/01/16/270653.htm'},
  {'d':'2025-05','st':'CA','co':'State Farm','k':'rate','ev':'Wins a 17% emergency interim rate increase post-fires (had sought 30%).','src':'https://calmatters.org/economy/2025/02/state-farm-insurance-rate-request-la-fires/'},
  {'d':'2025-11','st':'CA','co':'Farmers','k':'thaw','ev':'Lifts its new-policy cap — one of the first re-entry signals in California.','src':'https://matic.com/blog/2026-home-insurance-predictions/'},
  {'d':'2026-02','st':'LA','co':'Louisiana market','k':'thaw','ev':'LDI reports 2025 rate trends moderating to ~+4% after reform package.','src':'https://ldi.la.gov/news/press-releases/louisiana-insurance-market-shows-positive-rate-trends-in-2025-following-several-years-of-volatility'},
 ],
 'sources': [
  {'name':'U.S. Treasury FIO / NAIC — PCMI Supporting Underlying Metrics','vintage':'released Jan 2025 · data 2018–2022','role':'ZIP-level premiums/policy, nonrenewals, paid loss ratio, claim frequency & severity','url':'https://home.treasury.gov/news/press-releases/jy2791'},
  {'name':'U.S. Senate Budget Committee — county non-renewal data','vintage':'released Dec 2024 · data 2018–2023','role':'County non-renewal rates & policies in force (≈65% of national market)','url':'https://www.budget.senate.gov/chairman/newsroom/press/new-data-reveal-climate-change-driven-insurance-crisis-is-spreading'},
  {'name':'NOAA / NCEI — Billion-Dollar Weather & Climate Disasters','vintage':'final release · 1980–2024','role':'National disaster-cost ledger (CPI-adjusted). Product retired May 2025.','url':'https://www.ncei.noaa.gov/access/billions/'},
  {'name':'FEMA — OpenFEMA Disaster Declarations Summaries','vintage':'pulled Jul 2026 · 2000–present','role':'County-level major-disaster declarations by hazard type','url':'https://www.fema.gov/openfema-data-page/disaster-declarations-summaries-v2'},
  {'name':'U.S. Census — 2020 ZCTA ↔ county relationship file','vintage':'2020 vintage','role':'ZIP-to-county assignment (largest land-area overlap)','url':'https://www.census.gov/geographies/reference-files/time-series/geo/relationship-files.html'},
  {'name':'S&P Global Market Intelligence — approved HO rate changes','vintage':'via press reporting · 2023–2025','role':'MODELED nowcast anchors: +12.7% (2023), +10.4% (2024), +6.0% (2025)','url':'https://www.spglobal.com/market-intelligence/en/news-insights/articles/2025/1/us-homeowners-rates-rise-by-double-digits-for-2nd-straight-year-in-2024-87061085'},
  {'name':'us-atlas (U.S. Census cartography)','vintage':'counties-10m','role':'County & state topology','url':'https://github.com/topojson/us-atlas'},
 ],
}
with open(os.path.join(OUT,'national.js'),'w') as f:
    f.write('window.PM=window.PM||{};PM.national=')
    json.dump(nat_out, f, separators=(',',':'))
    f.write(';\n')

with open(os.path.join(OUT,'counties.js'),'w') as f:
    f.write('window.PM=window.PM||{};PM.counties=')
    json.dump(counties, f, separators=(',',':'))
    f.write(';\n')

topo = json.load(open(os.path.join(RAW,'counties-10m.json')))
with open(os.path.join(OUT,'topology.js'),'w') as f:
    f.write('window.PM=window.PM||{};PM.topology=')
    json.dump(topo, f, separators=(',',':'))
    f.write(';\n')

with open(os.path.join(OUT,'processed','county_metrics.csv'),'w',newline='') as f:
    w = csv.writer(f)
    w.writerow(['fips','county','state','prem2018','prem2019','prem2020','prem2021','prem2022',
                'prem2023_MODELED','prem2024_MODELED','prem2025_MODELED','prem_growth_pct_18_22',
                'nonrenew2018','nonrenew2019','nonrenew2020','nonrenew2021','nonrenew2022','nonrenew2023',
                'loss_ratio_avg','claim_freq_avg','decl_since2000','decl_2021plus','repricing_index_MODELED'])
    for g,o in sorted(counties.items()):
        p = o.get('p') or [None]*5; m = o.get('m') or [None]*3; nr = o.get('nr') or [None]*6
        lrs = [v for v in (o.get('lr') or []) if v is not None]
        cfs = [v for v in (o.get('cf') or []) if v is not None]
        w.writerow([g, o['n'], o['st'], *p, *m, o.get('gw'), *nr,
                    r2(sum(lrs)/len(lrs),3) if lrs else None,
                    r2(sum(cfs)/len(cfs),4) if cfs else None,
                    o.get('dr',{}).get('t'), o.get('dr',{}).get('l5'), o.get('dx')])
sizes = {fn: os.path.getsize(os.path.join(OUT,fn)) for fn in ('counties.js','national.js','topology.js')}
print('  sizes:', {k: f'{v/1e6:.2f}MB' for k,v in sizes.items()})
print('DONE')
